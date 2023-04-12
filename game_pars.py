import requests
from bs4 import BeautifulSoup
import time
import csv
import os
import openpyxl as op
from PIL import Image as pilIm
import openpyxl.drawing.image
from shutil import rmtree


def pars():
    save = []
    for i in range(1,20):
        url = f'https://stopgame.ru/games/filter?rating=izumitelno&p={i}'
        req = requests.get(url)
        src = req.text
        soup = BeautifulSoup(src,'lxml')
        all_games = soup.find_all(class_='item game-summary game-summary-horiz')
        for b,j in enumerate(all_games):
            try:
                image = j.find(class_ = 'image slanted').get('style')
                image_url = image.replace('background-image: url(','')[:-1]
                score = j.find(class_ = 'score').text.replace('\n',' ')
                name = j.find(class_ = 'caption caption-bold').find('a').text.replace('\n','')
                janr = j.find(class_= 'game-specs').find_all(class_ = 'game-spec')[1].find(class_ = 'value').find('a').text
            except:
                continue
            else:
                save.append([name,janr,score,image_url])
        time.sleep(1.3)
    return save

def to_csv():
    with open('best.csv', 'w', encoding='cp1251') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerow(['Название', 'Жанр', 'Оценка', 'Ссылка'])
    with open('best.csv', 'a', encoding='utf-8') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerows(pars())

def save_photo():
    save = pars()
    name = 'photos6'
    os.mkdir(f'{name}')
    for i in save:
        photo = requests.get(i[3])
        path = i[3].split('/')[-1]
        with open(f'C:\\Users\\alshe\\PycharmProjects\\parser2\\{name}\\{path}', 'wb') as ph:
            ph.write(photo.content)
    return save

def change_size(img,shirina):
    k = img.width/shirina
    img.width = shirina
    img.height = img.height/k
    return img

def exel():
    save = save_photo()
    path = os.getcwd()
    width = 150
    quality = 85
    os.mkdir('bad_images')
    photo_list = os.listdir(path+ '\photos6')
    for i in photo_list:
        image = pilIm.open(path + '\photos6' + f'\{i}')
        rgb_im = image.convert('RGB')
        rgb_im.save(path + '\\' + 'bad_images'+ '\\' + i,quality = quality)
    wb = op.Workbook()
    sheet = wb.active
    for v,i in enumerate(save,start=1):
        sheet.cell(row=v,column=1).value=i[0]
        sheet.cell(row=v,column=2).value=i[1]
        sheet.cell(row=v, column=3).value = i[2]
        img = op.drawing.image.Image('bad_images' + '\\' + i[3].split('/')[-1] )
        img = change_size(img, width)
        sheet.add_image(img, f'D{v}')
        sheet.row_dimensions[v].height = img.height * 0.75
    sheet.column_dimensions['D'].width = width*0.143
    wb.save('testoviy.xlsx')
    rmtree('bad_images')
    rmtree('photos6')

def main():
    exel()

if __name__ ==  '__main__':
    # photo_list = os.listdir('C:\\Users\\alshe\\PycharmProjects\\parser2\\'+'bad_images')
    # photo_list.sort()
    # print(photo_list)
    main()
