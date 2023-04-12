import requests
import time
from bs4 import BeautifulSoup
import openpyxl as op
import openpyxl.drawing.image
from os import mkdir
from shutil import rmtree

def pars():
    cookies = {
        'tmr_lvid': '4c5daabee74de68db76841a553432a16',
        'tmr_lvidTS': '1660651947568',
        '_ym_uid': '1660651948318749569',
        '_ym_d': '1660651948',
        'ouid': 'snyBDGL7ia1CfIH8gQyQAg==',
        '_gcl_au': '1.1.1204126964.1660651960',
        '_fbp': 'fb.1.1660651960754.1442523741',
        '_tt_enable_cookie': '1',
        '_ttp': '1534c905-4619-4b79-a6b2-f5e760eaf2eb',
        '__gads': 'ID=f72cfede7001797b:T=1660651959:S=ALNI_MZfUAPquDpAKC3ERMUHLW3r8rBusw',
        'ab_tracking_id': '6MjZL87k2sjCFFSQ',
        '_ga_SMLMFQCWFM': 'GS1.1.1660931892.1.0.1660931893.0.0.0',
        '_gaexp': 'GAX1.2.OLiZyqB6Q7GnGgoN6Iareg.19245.0',
        '_gid': 'GA1.2.434158453.1661683439',
        'tmr_detect': '0%7C1661683457616',
        'tmr_reqNum': '152',
        'referral': 'admitad%3Ade98015dd41761e77fb2724f4f3697fb',
        'catalog_session': 's8UWYy5JhxtAq5rUjayVZxrfgSZytY2b3xlatJeW',
        '_gat_UA-340679-1': '1',
        '_gat_UA-340679-16': '1',
        '__gpi': 'UID=00000acd81d275ba:T=1660651959:RT=1661713885:S=ALNI_MaXHlCRanSJ5swN9eRvbk88OOefrQ',
        '_ga_NG54S9EFTD': 'GS1.1.1661713885.26.1.1661713922.0.0.0',
        '_ga': 'GA1.1.247273602.1660651948',
        '_ga_4Y6NQKE48G': 'GS1.1.1661713885.26.1.1661713923.22.0.0',
    }

    headers = {
        'authority': 'catalog.onliner.by',
        'accept': 'application/json, text/javascript, */*; q=0.01',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'no-cache',
        # Requests sorts cookies= alphabetically
        # 'cookie': 'tmr_lvid=4c5daabee74de68db76841a553432a16; tmr_lvidTS=1660651947568; _ym_uid=1660651948318749569; _ym_d=1660651948; ouid=snyBDGL7ia1CfIH8gQyQAg==; _gcl_au=1.1.1204126964.1660651960; _fbp=fb.1.1660651960754.1442523741; _tt_enable_cookie=1; _ttp=1534c905-4619-4b79-a6b2-f5e760eaf2eb; __gads=ID=f72cfede7001797b:T=1660651959:S=ALNI_MZfUAPquDpAKC3ERMUHLW3r8rBusw; ab_tracking_id=6MjZL87k2sjCFFSQ; _ga_SMLMFQCWFM=GS1.1.1660931892.1.0.1660931893.0.0.0; _gaexp=GAX1.2.OLiZyqB6Q7GnGgoN6Iareg.19245.0; _gid=GA1.2.434158453.1661683439; tmr_detect=0%7C1661683457616; tmr_reqNum=152; referral=admitad%3Ade98015dd41761e77fb2724f4f3697fb; catalog_session=s8UWYy5JhxtAq5rUjayVZxrfgSZytY2b3xlatJeW; _gat_UA-340679-1=1; _gat_UA-340679-16=1; __gpi=UID=00000acd81d275ba:T=1660651959:RT=1661713885:S=ALNI_MaXHlCRanSJ5swN9eRvbk88OOefrQ; _ga_NG54S9EFTD=GS1.1.1661713885.26.1.1661713922.0.0.0; _ga=GA1.1.247273602.1660651948; _ga_4Y6NQKE48G=GS1.1.1661713885.26.1.1661713923.22.0.0',
        'pragma': 'no-cache',
        'referer': 'https://catalog.onliner.by/multicooker/redmond',
        'sec-ch-ua': '"Chromium";v="104", " Not A;Brand";v="99", "Google Chrome";v="104"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
    }
    names,min_prices,max_prices,image_url,urls,volume,coverage = [],[],[],[],[],[],[]
    response = requests.get('https://catalog.onliner.by/sdapi/catalog.api/search/multicooker?mfr[0]=redmond&group=1', cookies=cookies,headers=headers).json()
    page = response.get('page').get('last')
    for i in range(1,page+1):
        response = response.get('products')
        names += [ j.get('full_name') for j in response if j.get('prices') != None ]
        min_prices += [ j.get('prices').get('price_min').get('amount') for j in response if j.get('prices') != None ]
        max_prices += [j.get('prices').get('price_max').get('amount') for j in response if j.get('prices') != None]
        image_url += [ 'http:'+ j.get('images').get('header') for j in response if j.get('prices') != None ]
        urls += [j.get('html_url') for j in response if j.get('prices') != None]
        time.sleep(1)
        response = requests.get(f'https://catalog.onliner.by/sdapi/catalog.api/search/multicooker?mfr[0]=redmond&group=1&page={i}', cookies=cookies,headers=headers).json()
    for i in urls:
        response = requests.get(i,headers=headers)
        src = response.text
        soup = BeautifulSoup(src,'lxml')
        vol,cov =  soup.find(class_ = 'product-specs__main-group product-specs__group--full js-specs-full is-visible').find_all(class_ ='value__text')[4:6]
        volume.append(vol.text[0]+vol.text[-1])
        coverage.append(cov.text)
        time.sleep(0.2)
    wb = op.Workbook()
    sheet = wb.active
    mkdir('bad_images')
    for j,i in enumerate(image_url):
        photo = requests.get(i)
        name = i.split('/')[-1]
        with open('bad_images'+ '\\' + name,'wb') as ph:
            ph.write(photo.content)
        sheet.cell(row=j+1,column=1).value = names[j]
        sheet.cell(row=j + 1, column=3).value = min_prices[j]
        sheet.cell(row=j + 1, column=4).value = max_prices[j]
        sheet.cell(row=j + 1, column=5).value = volume[j]
        sheet.cell(row=j + 1, column=6).value = coverage[j]
        img = op.drawing.image.Image('bad_images'+ '\\' + name)
        sheet.add_image(img, f'B{j+1}')
        sheet.row_dimensions[j].height = img.height
        time.sleep(0.2)
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 25
    sheet.column_dimensions['F'].width = 25
    wb.save('s_perv_raza.xlsx')
    rmtree('bad_images')


def main():
    pars()

if __name__ == '__main__':
    main()